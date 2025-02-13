from openpyxl import load_workbook
import json

from Modelo import Veiculo


def ler_planilha(caminho: str) -> list[dict]:
    workbook = load_workbook(filename=caminho, data_only=True)

    nome_aba: str = workbook.sheetnames[0]
    workspace = workbook[nome_aba]

    cabecalhos: list[str] = [celula.value for celula in workspace[1]]

    dados: list[dict] = []
    for linha in workspace.iter_rows(min_row=2, values_only=True):
        linha_dict = dict(zip(cabecalhos, linha))
        if not linha_dict["Modelo"]:
            break
        dados.append(linha_dict)

    return dados

def salvar_json_saida(caminho: str) -> None:
    arquivo = open(caminho, mode="w", encoding="utf-8")
    json.dump(info_gasto_veiculos, arquivo, ensure_ascii=False, indent=4)
    arquivo.close()


caminho: str = "ARQUIVOS/Top10 Vendas.xlsx"
dados: list[dict] = ler_planilha(caminho=caminho)

info_gasto_veiculos: list[dict] = []
for dados_veiculo in dados:
    modelo: str = dados_veiculo["Modelo"]
    if "*" in modelo:
        continue
    if modelo == "Creta":
        combustao: bool = True
    else:
        combustao = False
    veiculo = Veiculo(dados_veiculo=dados_veiculo, combustao=combustao)
    veiculo.processar()
    gastos_por_km: float = veiculo.retornar_gasto_total_por_km()
    gastos_detalhados: dict = veiculo.retornar_gastos_detalhados()
    info_veiculo = {
        "Modelo": modelo,
        "Gasto por Km": gastos_por_km,
        "Gastos Detalhados": gastos_detalhados,
    }
    info_gasto_veiculos.append(info_veiculo)

caminho_saida: str = "info.json"
salvar_json_saida(caminho=caminho_saida)
